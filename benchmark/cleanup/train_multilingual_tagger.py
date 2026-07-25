#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.14,<3.15"
# dependencies = [
#   "numpy>=2.3,<3",
#   "onnx>=1.20,<2",
#   "onnxruntime>=1.27,<2",
#   "openpyxl>=3.1,<4",
#   "torch>=2.13,<3",
#   "transformers>=5.4,<6",
# ]
# ///
"""Train and export a deletion-only multilingual disfluency tagger.

The input is the DISCO repository's ``data/labeled-data`` directory. The
exported model never generates text: it predicts KEEP or DELETE for each input
word. Runtime reconstruction therefore remains restricted to original tokens.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import random
import re
import time
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

KEEP = 0
DELETE = 1
IGNORE = -100
LANGUAGE_FILES = {
    "de": "German.xlsx",
    "en": "English.xlsx",
    "fr": "French.xlsx",
    "hi": "Hindi.xlsx",
}
_SPANS = re.compile(r"\S+")


@dataclass(frozen=True, slots=True)
class Example:
    words: tuple[str, ...]
    labels: tuple[int, ...]
    language: str


@dataclass(frozen=True, slots=True)
class ThresholdMetrics:
    threshold: float
    delete_precision: float
    delete_recall: float
    delete_f1: float
    false_delete_rate: float
    accuracy: float
    true_deletes: int
    false_deletes: int
    missed_deletes: int
    true_keeps: int


def normalized_word(value: str) -> str:
    """Return the comparison form used to align fluent text to source words."""
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return "".join(
        character
        for character in normalized
        if not unicodedata.category(character).startswith(("P", "Z"))
    )


def derive_example(disfluent: str, fluent: str, language: str) -> Example | None:
    """Label source words by matching the fluent subsequence from right to left."""
    words = tuple(match.group() for match in _SPANS.finditer(disfluent.strip()))
    fluent_words = tuple(match.group() for match in _SPANS.finditer(fluent.strip()))
    if not words:
        return None

    source_keys = tuple(normalized_word(word) for word in words)
    target_keys = tuple(key for word in fluent_words if (key := normalized_word(word)))
    labels = [DELETE] * len(words)
    target_index = len(target_keys) - 1
    for source_index in range(len(words) - 1, -1, -1):
        source_key = source_keys[source_index]
        if source_key and target_index >= 0 and source_key == target_keys[target_index]:
            labels[source_index] = KEEP
            target_index -= 1

    if target_index >= 0:
        return None
    return Example(words=words, labels=tuple(labels), language=language)


def _find_column(headers: tuple[str, ...], kind: str) -> int:
    for index, header in enumerate(headers):
        normalized = " ".join(header.casefold().split())
        if kind == "disfluent" and "disfluent sentence" in normalized:
            return index
        if (
            kind == "fluent"
            and "fluent sentence" in normalized
            and "disfluent sentence" not in normalized
        ):
            return index
    raise ValueError(f"workbook does not contain a {kind!r} sentence column")


def load_workbook_examples(path: Path, language: str) -> tuple[list[Example], int]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = tuple("" if value is None else str(value) for value in raw_headers)
    disfluent_column = _find_column(headers, "disfluent")
    fluent_column = _find_column(headers, "fluent")

    examples: list[Example] = []
    rejected = 0
    for row in rows:
        disfluent_value = row[disfluent_column]
        fluent_value = row[fluent_column]
        if disfluent_value is None or fluent_value is None:
            continue
        example = derive_example(str(disfluent_value), str(fluent_value), language)
        if example is None:
            rejected += 1
        else:
            examples.append(example)
    workbook.close()
    return examples, rejected


def load_disco(data_dir: Path) -> tuple[list[Example], dict[str, Any]]:
    examples: list[Example] = []
    accepted_by_language: dict[str, int] = {}
    rejected_by_language: dict[str, int] = {}
    for language, filename in LANGUAGE_FILES.items():
        language_examples, rejected = load_workbook_examples(data_dir / filename, language)
        examples.extend(language_examples)
        accepted_by_language[language] = len(language_examples)
        rejected_by_language[language] = rejected
    token_count = sum(len(example.labels) for example in examples)
    delete_count = sum(sum(example.labels) for example in examples)
    delete_ratio = delete_count / max(1, token_count)
    if not 0.1 <= delete_ratio <= 0.35:
        raise ValueError(
            f"unexpected DISCO label density: {delete_ratio:.2%} DELETE labels; expected 10% to 35%"
        )
    return examples, {
        "accepted_by_language": accepted_by_language,
        "rejected_by_language": rejected_by_language,
        "accepted_total": len(examples),
        "rejected_total": sum(rejected_by_language.values()),
        "token_count": token_count,
        "delete_count": delete_count,
        "delete_ratio": delete_ratio,
    }


def split_examples(
    examples: list[Example],
    *,
    seed: int,
) -> tuple[list[Example], list[Example], list[Example]]:
    grouped: dict[str, list[Example]] = {}
    for example in examples:
        grouped.setdefault(example.language, []).append(example)

    training: list[Example] = []
    validation: list[Example] = []
    testing: list[Example] = []
    for language, language_examples in sorted(grouped.items()):
        shuffled = list(language_examples)
        random.Random(f"{seed}:{language}").shuffle(shuffled)
        validation_start = math.floor(len(shuffled) * 0.8)
        testing_start = math.floor(len(shuffled) * 0.9)
        training.extend(shuffled[:validation_start])
        validation.extend(shuffled[validation_start:testing_start])
        testing.extend(shuffled[testing_start:])

    random.Random(seed).shuffle(training)
    random.Random(seed + 1).shuffle(validation)
    random.Random(seed + 2).shuffle(testing)
    return training, validation, testing


def metrics_for_threshold(
    probabilities: list[float],
    labels: list[int],
    threshold: float,
) -> ThresholdMetrics:
    true_deletes = false_deletes = missed_deletes = true_keeps = 0
    for probability, label in zip(probabilities, labels, strict=True):
        predicted_delete = probability >= threshold
        if predicted_delete and label == DELETE:
            true_deletes += 1
        elif predicted_delete:
            false_deletes += 1
        elif label == DELETE:
            missed_deletes += 1
        else:
            true_keeps += 1

    precision = true_deletes / max(1, true_deletes + false_deletes)
    recall = true_deletes / max(1, true_deletes + missed_deletes)
    f1 = 2 * precision * recall / max(1e-12, precision + recall)
    false_delete_rate = false_deletes / max(1, false_deletes + true_keeps)
    accuracy = (true_deletes + true_keeps) / max(1, len(labels))
    return ThresholdMetrics(
        threshold=threshold,
        delete_precision=precision,
        delete_recall=recall,
        delete_f1=f1,
        false_delete_rate=false_delete_rate,
        accuracy=accuracy,
        true_deletes=true_deletes,
        false_deletes=false_deletes,
        missed_deletes=missed_deletes,
        true_keeps=true_keeps,
    )


def choose_threshold(
    probabilities: list[float],
    labels: list[int],
    *,
    max_false_delete_rate: float,
) -> tuple[float, list[ThresholdMetrics]]:
    candidates = [
        0.5,
        0.6,
        0.7,
        0.75,
        0.8,
        0.85,
        0.9,
        0.925,
        0.95,
        0.975,
        0.99,
        0.995,
    ]
    metrics = [metrics_for_threshold(probabilities, labels, threshold) for threshold in candidates]
    eligible = [item for item in metrics if item.false_delete_rate <= max_false_delete_rate]
    selected = max(
        eligible or metrics,
        key=lambda item: (item.delete_f1, item.delete_precision, item.threshold),
    )
    return selected.threshold, metrics


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--data-dir",
        type=Path,
        required=True,
        help="DISCO data/labeled-data directory",
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--base-model",
        default="distilbert/distilbert-base-multilingual-cased",
    )
    parser.add_argument("--epochs", type=int, default=3)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--learning-rate", type=float, default=3e-5)
    parser.add_argument("--max-length", type=int, default=128)
    parser.add_argument("--seed", type=int, default=1337)
    parser.add_argument("--threads", type=int, default=min(8, os.cpu_count() or 1))
    parser.add_argument("--device", choices=("auto", "cpu", "cuda"), default="auto")
    parser.add_argument("--max-false-delete-rate", type=float, default=0.0025)
    return parser.parse_args()


def main() -> None:
    import numpy as np
    import onnx
    import onnxruntime as ort
    import torch
    from onnxruntime.quantization import QuantType, quantize_dynamic
    from torch.utils.data import DataLoader, TensorDataset
    from transformers import (
        AutoModelForTokenClassification,
        AutoTokenizer,
        get_linear_schedule_with_warmup,
    )

    args = parse_args()
    if args.epochs <= 0 or args.batch_size <= 0 or args.max_length < 16:
        raise ValueError("epochs and batch size must be positive; max length must be >= 16")
    args.output.mkdir(parents=True, exist_ok=True)
    torch.manual_seed(args.seed)
    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.set_num_threads(args.threads)
    torch.set_num_interop_threads(1)

    if args.device == "cuda" and not torch.cuda.is_available():
        raise RuntimeError("CUDA was requested but is unavailable")
    device = torch.device(
        "cuda"
        if args.device == "cuda" or (args.device == "auto" and torch.cuda.is_available())
        else "cpu"
    )

    load_started = time.monotonic()
    examples, data_stats = load_disco(args.data_dir)
    training, validation, testing = split_examples(examples, seed=args.seed)
    print(
        f"loaded {len(examples)} examples "
        f"(train={len(training)}, val={len(validation)}, test={len(testing)})"
    )
    print(json.dumps(data_stats, ensure_ascii=False))

    tokenizer = AutoTokenizer.from_pretrained(args.base_model, use_fast=True)

    def encode(items: list[Example]) -> TensorDataset:
        encoded = tokenizer(
            [list(example.words) for example in items],
            is_split_into_words=True,
            max_length=args.max_length,
            padding="max_length",
            truncation=True,
            return_tensors="pt",
        )
        aligned_labels: list[list[int]] = []
        for batch_index, example in enumerate(items):
            word_ids = encoded.word_ids(batch_index=batch_index)
            previous_word_id: int | None = None
            item_labels: list[int] = []
            for word_id in word_ids:
                if word_id is None or word_id == previous_word_id:
                    item_labels.append(IGNORE)
                else:
                    item_labels.append(example.labels[word_id])
                previous_word_id = word_id
            aligned_labels.append(item_labels)
        return TensorDataset(
            encoded["input_ids"],
            encoded["attention_mask"],
            torch.tensor(aligned_labels, dtype=torch.long),
        )

    train_dataset = encode(training)
    validation_dataset = encode(validation)
    test_dataset = encode(testing)
    train_loader = DataLoader(
        train_dataset,
        batch_size=args.batch_size,
        shuffle=True,
    )
    validation_loader = DataLoader(
        validation_dataset,
        batch_size=args.batch_size,
        shuffle=False,
    )
    test_loader = DataLoader(
        test_dataset,
        batch_size=args.batch_size,
        shuffle=False,
    )

    model = AutoModelForTokenClassification.from_pretrained(
        args.base_model,
        num_labels=2,
        id2label={KEEP: "KEEP", DELETE: "DELETE"},
        label2id={"KEEP": KEEP, "DELETE": DELETE},
    ).to(device)
    optimizer = torch.optim.AdamW(
        model.parameters(),
        lr=args.learning_rate,
        weight_decay=0.01,
    )
    total_steps = len(train_loader) * args.epochs
    scheduler = get_linear_schedule_with_warmup(
        optimizer,
        num_warmup_steps=max(1, round(total_steps * 0.1)),
        num_training_steps=total_steps,
    )

    def evaluate(loader: DataLoader[tuple[torch.Tensor, ...]]) -> tuple[list[float], list[int]]:
        model.eval()
        probabilities: list[float] = []
        labels: list[int] = []
        with torch.inference_mode():
            for input_ids, attention_mask, batch_labels in loader:
                logits = model(
                    input_ids=input_ids.to(device),
                    attention_mask=attention_mask.to(device),
                ).logits
                delete_probabilities = torch.softmax(logits, dim=-1)[..., DELETE]
                valid = batch_labels != IGNORE
                probabilities.extend(delete_probabilities.cpu()[valid].tolist())
                labels.extend(batch_labels[valid].tolist())
        return probabilities, labels

    checkpoint_dir = args.output / "checkpoint"
    best_validation_f1 = -1.0
    training_started = time.monotonic()
    for epoch in range(1, args.epochs + 1):
        model.train()
        total_loss = 0.0
        for step, (input_ids, attention_mask, labels) in enumerate(train_loader, start=1):
            optimizer.zero_grad(set_to_none=True)
            result = model(
                input_ids=input_ids.to(device),
                attention_mask=attention_mask.to(device),
                labels=labels.to(device),
            )
            result.loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
            optimizer.step()
            scheduler.step()
            total_loss += float(result.loss.detach().cpu())
            if step % 50 == 0 or step == len(train_loader):
                print(
                    f"epoch={epoch}/{args.epochs} step={step}/{len(train_loader)} "
                    f"loss={total_loss / step:.4f}",
                    flush=True,
                )

        validation_probabilities, validation_labels = evaluate(validation_loader)
        validation_metrics = metrics_for_threshold(
            validation_probabilities,
            validation_labels,
            0.5,
        )
        print(
            f"epoch={epoch} validation delete_f1={validation_metrics.delete_f1:.4f} "
            f"precision={validation_metrics.delete_precision:.4f} "
            f"recall={validation_metrics.delete_recall:.4f}",
            flush=True,
        )
        if validation_metrics.delete_f1 > best_validation_f1:
            best_validation_f1 = validation_metrics.delete_f1
            model.save_pretrained(checkpoint_dir)
            tokenizer.save_pretrained(checkpoint_dir)

    model = AutoModelForTokenClassification.from_pretrained(checkpoint_dir).to(device)
    validation_probabilities, validation_labels = evaluate(validation_loader)
    selected_threshold, threshold_metrics = choose_threshold(
        validation_probabilities,
        validation_labels,
        max_false_delete_rate=args.max_false_delete_rate,
    )
    test_probabilities, test_labels = evaluate(test_loader)
    test_metrics = metrics_for_threshold(
        test_probabilities,
        test_labels,
        selected_threshold,
    )
    print(
        f"selected threshold={selected_threshold:.3f}; "
        f"test delete_f1={test_metrics.delete_f1:.4f} "
        f"precision={test_metrics.delete_precision:.4f} "
        f"recall={test_metrics.delete_recall:.4f} "
        f"false_delete_rate={test_metrics.false_delete_rate:.5f}"
    )

    tokenizer.save_pretrained(args.output)
    model = model.to("cpu").eval()

    class LogitsOnly(torch.nn.Module):
        def __init__(self, wrapped: torch.nn.Module) -> None:
            super().__init__()
            self.wrapped = wrapped

        def forward(
            self,
            input_ids: torch.Tensor,
            attention_mask: torch.Tensor,
        ) -> torch.Tensor:
            return self.wrapped(
                input_ids=input_ids,
                attention_mask=attention_mask,
            ).logits

    dummy_input_ids = torch.ones((1, args.max_length), dtype=torch.long)
    dummy_attention_mask = torch.ones((1, args.max_length), dtype=torch.long)
    fp32_path = args.output / "model.fp32.onnx"
    int8_path = args.output / "model.int8.onnx"
    torch.onnx.export(
        LogitsOnly(model),
        (dummy_input_ids, dummy_attention_mask),
        fp32_path,
        input_names=["input_ids", "attention_mask"],
        output_names=["logits"],
        dynamic_axes={
            "input_ids": {0: "batch", 1: "sequence"},
            "attention_mask": {0: "batch", 1: "sequence"},
            "logits": {0: "batch", 1: "sequence"},
        },
        do_constant_folding=True,
        dynamo=False,
        opset_version=17,
    )
    onnx.checker.check_model(onnx.load(fp32_path))
    quantize_dynamic(
        fp32_path,
        int8_path,
        per_channel=True,
        weight_type=QuantType.QInt8,
    )
    onnx.checker.check_model(onnx.load(int8_path))

    session = ort.InferenceSession(
        str(int8_path),
        providers=["CPUExecutionProvider"],
    )
    probe = tokenizer(
        "um we we should ship this tomorrow",
        max_length=args.max_length,
        padding="max_length",
        truncation=True,
        return_tensors="np",
    )
    session.run(
        None,
        {
            "input_ids": probe["input_ids"].astype(np.int64),
            "attention_mask": probe["attention_mask"].astype(np.int64),
        },
    )

    metadata = {
        "model_id": "multilingual-edit-tagger",
        "architecture": "distilbert-token-classification",
        "base_model": args.base_model,
        "labels": {"KEEP": KEEP, "DELETE": DELETE},
        "languages": sorted(LANGUAGE_FILES),
        "delete_threshold": selected_threshold,
        "max_length": args.max_length,
        "max_delete_ratio": 0.35,
        "training": {
            "epochs": args.epochs,
            "batch_size": args.batch_size,
            "learning_rate": args.learning_rate,
            "seed": args.seed,
            "device": str(device),
            "seconds": round(time.monotonic() - training_started, 1),
            "data": data_stats,
            "split_sizes": {
                "training": len(training),
                "validation": len(validation),
                "testing": len(testing),
            },
            "validation_thresholds": [asdict(item) for item in threshold_metrics],
            "test": asdict(test_metrics),
        },
        "source": {
            "dataset": "DISCO",
            "repository": "https://github.com/vineet2104/DISCO",
            "license_note": (
                "The upstream repository did not contain an explicit LICENSE file "
                "when this experiment was run; do not redistribute the trained "
                "artifact until dataset/model licensing is confirmed."
            ),
        },
        "total_seconds": round(time.monotonic() - load_started, 1),
    }
    (args.output / "metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n"
    )
    print(
        json.dumps(
            {
                "artifact": str(int8_path),
                "artifact_mib": round(int8_path.stat().st_size / 1024 / 1024, 1),
                "test": asdict(test_metrics),
                "delete_threshold": selected_threshold,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
